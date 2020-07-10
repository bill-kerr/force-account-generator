from openpyxl import load_workbook as load_wb
from worksheet import Worksheet
from config import Config


class Workbook:
    def __init__(self, excel_file_path):
        try:
            self.__load_workbook(excel_file_path)
        except Exception:
            raise Exception('Error loading workbook.')

        self.__generate_defined_cells()
        self.__generate_worksheets()

    def __load_workbook(self, file_path):
        self.__workbook = load_wb(file_path, read_only=True, data_only=True)

    def __generate_defined_cells(self):
        cell_names = Config.defined_cells['defined_cells'].values()
        self.defined_cells = self.__get_defined_cells_from_workbook(cell_names)

    def __get_defined_cells_from_workbook(self, cell_names):
        defined_cells = {sheet_name: {} for sheet_name in self.__workbook.get_sheet_names()}
        for cell_name in cell_names:
            for worksheet_name, cell_coordinates in self.__workbook.defined_names[cell_name].destinations:
                worksheet = self.__workbook[worksheet_name]
                defined_cells[worksheet_name][cell_name] = worksheet[cell_coordinates].value

        return defined_cells

    def __generate_worksheets(self):
        sheet_names = config['worksheet_names']
        self.worksheets = {}

        for name in sheet_names:
            sheet_name = sheet_names[name]
            self.worksheets[name] = Worksheet(self.__workbook[sheet_name], self.defined_cells[sheet_name])

    def get_named_cell_values(self, cell_names=config['defined_cells']):
        return [self.get_named_cell_value(cell_name) for cell_name in cell_names]
