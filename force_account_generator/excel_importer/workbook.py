from openpyxl import load_workbook as load_wb
from .config import Config
from .worksheet import Worksheet
from .material import process_material
from .labor import process_labor
from .equipment import process_equipment
from .consumables import process_consumables
from .rentals_and_services import process_rentals_and_services


class Workbook:
    def __init__(self, source_file):
        self.__load_workbook(source_file)
        self.__sheet_map = {name: key for key, name in Config.worksheet_names.items()}
        self.__defined_cell_map = {name: key for key, name in Config.defined_cells.items()}
        self.__generate_defined_cells()
        self.__generate_worksheets()

    def __load_workbook(self, source_file):
        try:
            self._workbook = load_wb(source_file, read_only=True, data_only=True)
        except Exception:
            raise Exception('Error loading workbook from ' + source_file)

    def __generate_defined_cells(self):
        cell_names = Config.defined_cells.values()
        self.defined_cells = {name: {} for name in Config.worksheet_names}
        for cell_name in cell_names:
            for worksheet_name, cell_coordinates in self.__workbook.defined_names[cell_name].destinations:
                sheet_name = self._sheet_map[worksheet_name]
                defined_cell = self.__defined_cell_map[cell_name]
                self.defined_cells[sheet_name][defined_cell] = self.__workbook[worksheet_name][cell_coordinates].value

    def __generate_worksheets(self):
        self.worksheets = {}
        for name, org_name in Config.worksheet_names.items():
            process_function = self.__get_process_function(name)
            worksheet = self.__workbook[org_name]
            defined_cells = self.defined_cells[name]
            self.worksheets[name] = Worksheet(worksheet, defined_cells, process_function)

    def __get_process_function(self, sheet_name):
        process_functions = {
            'material': process_material,
            'labor': process_labor,
            'equipment': process_equipment,
            'rentals_and_services': process_rentals_and_services,
            'consumables': process_consumables
        }
        return process_functions.get(sheet_name)

    def __get_summary_cells(self):
        return {
            'county': self.defined_cells['summary'].get('county'),
            'state_route': self.defined_cells['summary'].get('state_route'),
            'section': self.defined_cells['summary'].get('section'),
            'work_order_number': self.defined_cells['summary'].get('work_order_number'),
            'contract': self.defined_cells['summary'].get('contract'),
            'item_number': self.defined_cells['summary'].get('item_number'),
            'prime_contractor': self.defined_cells['summary'].get('prime_contractor'),
            'statement_of_cost': self.defined_cells['summary'].get('statement_of_cost')
        }

    def __process_worksheets(self):
        data = {}
        for i, worksheet in enumerate(self.worksheets.values()):
            data.update(worksheet.process())
        return data

    def process_workbook(self):
        data = {}
        data.update(self.__get_summary_cells())
        data.update(self.__process_worksheets())
        return data
